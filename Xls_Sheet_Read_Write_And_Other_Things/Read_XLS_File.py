import openpyxl
from openpyxl import Workbook


def Read_xls_File():
    book = Workbook()
    sheet = book.active

    # sheet_obj = wb_obj.active
    # cell_obj = sheet_obj.cell(row=6, column=1)
    # print(cell_obj.value)  # Select only One Row or column which you want to show
    # source_list = []
    # number_of_Rows = sheet_obj.max_row

    # for i in range(3, number_of_Rows + 1):
    #     row = sheet_obj.cell(row=i, column=1)
    #     row2 = sheet_obj.append(row='test', column=2)
    #     # print(row.value)  # Print all rows values
    #     # print(row2.value)
    #     source_list.append(row.value+" - "+row2.value)
    #     print(row.value+" - "+row2.value)
    #     print(source_list)
    number_of_Rows = sheet_obj.max_row

    book.save('appending.xlsx')
        
Read_xls_File()
